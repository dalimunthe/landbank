import math
from collections import Counter
from itertools import groupby

import arcpy
import datetime

from openpyxl import load_workbook, Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, Alignment
from openpyxl.styles.borders import Border, Side
from reportlab.lib import colors
from reportlab.lib.colors import (
    black
)
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm, mm
from reportlab.pdfgen import canvas
from reportlab.platypus import BaseDocTemplate, PageTemplate, Paragraph, Spacer, TableStyle, Frame, Table
from reportlab.platypus.flowables import HRFlowable

fc = r"C:\Users\Rasyid GIS\Documents\LANDBANK\ptw\PYTW.shp"
fcs = r"C:\Users\Rasyid GIS\Documents\LANDBANK\data\JPTIBP0009IJL3.shp"
fcss = r"C:\Users\Rasyid GIS\Documents\LANDBANK\data\NPTAJP0009IJL3.shp"


class convertVertices(object):
    def __init__(self,
                 feature_class,
                 legal_type="CAD",
                 header=False,
                 rounding=2,
                 nama_pt=''):
        self.feature_class = feature_class
        area_of_field = []
        for row in arcpy.da.SearchCursor(self.feature_class, ["SHAPE@AREA"]):
            area_of_field.append(round(row[0] / 10000, 2))
        self.legal_type = legal_type
        self.header = header
        self.rounding = rounding
        self.nama_pt = nama_pt
        self.area_ha = '{:,.2f} ha'.format(area_of_field[0])
        self.feature_describe = arcpy.Describe(self.feature_class)
        self.srcode = self.feature_describe.spatialReference.name.replace("_", " ")
        self.wkid = self.feature_describe.spatialReference.factoryCode
        self.proyeksi = str(self.srcode) + " (" + str(self.wkid) + ")"

    def information(self):
        years = int(datetime.date.today().strftime('%Y'))
        months = int(datetime.date.today().strftime('%m'))
        days = int(datetime.date.today().strftime('%d'))
        info = [['Landbank ID : ', ''],
                ['Nama PT : ', self.nama_pt],
                ['Luas : ', self.area_ha],
                ['Tanggal : ', datetime.datetime(years, months, days)],
                ['Proyeksi : ', self.proyeksi]
                ]
        return info

    def explode_polygon(self):
        dum_row = []
        rows_explode = arcpy.da.SearchCursor(self.feature_class, ["SHAPE@", "No_SK", "Remarks"])
        skiplastvertex = True
        for partno, row in enumerate(rows_explode):
            for partindex, part in enumerate(row[0]):
                if row[0].isMultipart:
                    polygonno = partindex + 1
                else:
                    polygonno = partno + 1
                ringindex = 0
                vertexindex = 0 + 1
                fid = 0
                no_sk = row[1]
                pnt = part.next()
                while pnt:
                    output = [pnt.X, pnt.Y, fid, partindex, ringindex, vertexindex, no_sk]
                    if ringindex == 0:
                        ringStatus = "in"
                    else:
                        ringStatus = "out"
                    pnt = part.next()
                    if pnt is None:
                        if not skiplastvertex:
                            dum_row.append(
                                [polygonno, ringStatus, vertexindex,
                                 round(pnt.X, self.rounding), round(pnt.Y, self.rounding), output[6]])
                        pnt = part.next()
                        if pnt:
                            vertexindex = 0 + 1
                            ringindex += 1
                    else:
                        dum_row.append(
                            [polygonno, ringStatus, vertexindex,
                             round(pnt.X, self.rounding), round(pnt.Y, self.rounding), output[6]])
                        vertexindex += 1

        raw_distance = [0]
        for idx in range(1, len(dum_row)):
            loc1 = dum_row[idx - 1]
            loc2 = dum_row[idx]

            lat1 = loc1[3]
            lon1 = loc1[4]

            lat2 = loc2[3]
            lon2 = loc2[4]

            first_point = math.pow((lat1 - lat2), self.rounding)
            second_point = math.pow((lon1 - lon2), self.rounding)

            dis = round(float(math.sqrt(first_point + second_point)), self.rounding)
            raw_distance.append(dis)
        total = 0

        x_val = Counter([x for (p_no, r_s, v_i, x, y, n_s) in dum_row])
        y_val = Counter([y for (p_no, r_s, v_i, x, y, n_s) in dum_row])
        x_dup = []
        y_dup = []
        for idx, key in enumerate(x_val):
            if x_val[key] > 1:
                x_dup.append(key)
            else:
                pass
        for idx, key in enumerate(y_val):
            if y_val[key] > 1:
                y_dup.append(key)
            else:
                pass
        if self.header == True:
            final_arr = [["Polygon No.",
                          "In/Out",
                          "Point Order",
                          "Easting (X)",
                          "Northing (Y)",
                          "Nomor SK",
                          "Distance",
                          "Shared Vertices",
                          "Legal Type",
                          "Point ID",
                          "No"]]
        else:
            final_arr = []

        for idx, ax in enumerate(dum_row):
            if ax[2] == 1:
                distances = 0
            else:
                distances = raw_distance[idx]
            if ax[3] in x_dup and ax[4] in y_dup:
                l = [dum_row.index(i) + 1 for i in dum_row if i[3] == ax[3] and ax[4] == i[4] in i]
                ind = []
                ind.append(idx + 1)
                ss = set(l) - set(ind)
                aa = list(ss)
                ddd = ', '.join(str(e) for e in aa)
                all_match = ', '.join(str(e) for e in l)
                final_arr.append([ax[0], ax[1],
                                  ax[2], ax[3],
                                  ax[4], ax[5],
                                  distances, "SHARED (" + str(all_match) + ")", self.legal_type, "", idx + 1])
            else:
                final_arr.append([ax[0], ax[1],
                                  ax[2], ax[3],
                                  ax[4], ax[5],
                                  distances, "SINGLE", self.legal_type, "", idx + 1])
        return final_arr


class exportExcel:
    def __init__(self,
                 path='',
                 file_name='',
                 info_array=[],
                 data_array=[]):
        self.path = path + "//"
        self.file_name = file_name + ".xlsx"
        self.info_array = info_array
        self.data_array = data_array

    def round_trailing(self, number):
        return '{:,.2f}'.format(number)

    def createFile(self):
        wb = Workbook()
        sheet = wb.active
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        thin_border_left_right = Border(left=Side(style='thin'),
                                        right=Side(style='thin'))
        thin_border_bottom_left = Border(left=Side(style='thin'),
                                         bottom=Side(style='thin'))
        thin_border_bottom_left_right = Border(left=Side(style='thin'),
                                               bottom=Side(style='thin'),
                                               right=Side(style='thin'))
        header_alignement = Alignment(vertical='center', horizontal='center', wrapText=True)
        header_font = Font(size=11, bold=True)

        comment_s = Comment("Mandatory Field\nEx: JPTIBP0009IJL3", "GIS")
        comment_d = Comment("Mandatory Field\nEx: PT. Intitama Berlian Perkebunan", "GIS")
        comment_t = Comment("Mandatory Field\nEx: BPN001", "GIS")
        comment_e = Comment("Mandatory Field\nEx: IJL", "GIS")
        sheet["C2"].comment = comment_s
        sheet["C3"].comment = comment_d
        sheet["B8"].comment = comment_t
        sheet["g8"].comment = comment_e

        for row, i in enumerate(self.info_array):
            sheet['A' + str(row + 2)] = i[0]
            sheet['C' + str(row + 2)] = i[1]

            sheet['A' + str(row + 2)].font = header_font
            sheet['C' + str(row + 2)].font = header_font

        cell = sheet['C5']
        cell.number_format = 'DD/MM/YYYY'
        cell.alignment = Alignment(horizontal='left')

        sheet.merge_cells('A2:B2')
        sheet.merge_cells('A3:B3')
        sheet.merge_cells('A4:B4')
        sheet.merge_cells('A5:B5')
        sheet.merge_cells('A6:B6')

        for row, i in enumerate(self.data_array):
            # WRITE
            sheet['A' + str(row + 8)] = i[10]
            sheet['B' + str(row + 8)] = i[9]
            sheet['C' + str(row + 8)] = i[3]
            sheet['D' + str(row + 8)] = i[4]
            sheet['E' + str(row + 8)] = i[6]
            sheet['F' + str(row + 8)] = i[7]
            sheet['G' + str(row + 8)] = i[8]
            sheet['H' + str(row + 8)] = i[0]
            sheet['I' + str(row + 8)] = i[1]
            sheet['J' + str(row + 8)] = i[2]
            sheet['K' + str(row + 8)] = i[5]

            # STYLE
            sheet['A' + str(row + 8)].border = thin_border_left_right
            sheet['A' + str(row + 8)].border = thin_border_left_right
            sheet['B' + str(row + 8)].border = thin_border_left_right
            sheet['C' + str(row + 8)].border = thin_border_left_right
            sheet['D' + str(row + 8)].border = thin_border_left_right
            sheet['E' + str(row + 8)].border = thin_border_left_right
            sheet['F' + str(row + 8)].border = thin_border_left_right
            sheet['G' + str(row + 8)].border = thin_border_left_right
            sheet['H' + str(row + 8)].border = thin_border_left_right
            sheet['I' + str(row + 8)].border = thin_border_left_right
            sheet['J' + str(row + 8)].border = thin_border_left_right
            sheet['K' + str(row + 8)].border = thin_border_left_right

            # HEADER
            if row == 0:
                # STYLE
                sheet['A' + str(row + 8)].font = header_font
                sheet['B' + str(row + 8)].font = header_font
                sheet['C' + str(row + 8)].font = header_font
                sheet['D' + str(row + 8)].font = header_font
                sheet['E' + str(row + 8)].font = header_font
                sheet['F' + str(row + 8)].font = header_font
                sheet['G' + str(row + 8)].font = header_font
                sheet['H' + str(row + 8)].font = header_font
                sheet['I' + str(row + 8)].font = header_font
                sheet['J' + str(row + 8)].font = header_font
                sheet['K' + str(row + 8)].font = header_font

                sheet['A' + str(row + 8)].alignment = header_alignement
                sheet['B' + str(row + 8)].alignment = header_alignement
                sheet['C' + str(row + 8)].alignment = header_alignement
                sheet['D' + str(row + 8)].alignment = header_alignement
                sheet['E' + str(row + 8)].alignment = header_alignement
                sheet['F' + str(row + 8)].alignment = header_alignement
                sheet['G' + str(row + 8)].alignment = header_alignement
                sheet['H' + str(row + 8)].alignment = header_alignement
                sheet['I' + str(row + 8)].alignment = header_alignement
                sheet['J' + str(row + 8)].alignment = header_alignement
                sheet['K' + str(row + 8)].alignment = header_alignement

                sheet['A' + str(row + 8)].border = thin_border
                sheet['A' + str(row + 8)].border = thin_border
                sheet['B' + str(row + 8)].border = thin_border
                sheet['C' + str(row + 8)].border = thin_border
                sheet['D' + str(row + 8)].border = thin_border
                sheet['E' + str(row + 8)].border = thin_border
                sheet['F' + str(row + 8)].border = thin_border
                sheet['G' + str(row + 8)].border = thin_border
                sheet['H' + str(row + 8)].border = thin_border
                sheet['I' + str(row + 8)].border = thin_border
                sheet['J' + str(row + 8)].border = thin_border
                sheet['K' + str(row + 8)].border = thin_border
                # BODY
            elif row > 0 and row < (len(self.data_array) - 1):
                # WRITE

                # STYLE
                sheet['C' + str(row + 8)].number_format = '#,##0.00'
                sheet['D' + str(row + 8)].number_format = '#,##0.00'
                sheet['E' + str(row + 8)].number_format = '#,##0.00'
            # FOOTER
            else:
                # WRITE

                # STYLE
                sheet['C' + str(row + 8)].number_format = '#,##0.00'
                sheet['D' + str(row + 8)].number_format = '#,##0.00'
                sheet['E' + str(row + 8)].number_format = '#,##0.00'

                sheet['A' + str(row + 8)].border = thin_border_bottom_left
                sheet['A' + str(row + 8)].border = thin_border_bottom_left
                sheet['B' + str(row + 8)].border = thin_border_bottom_left
                sheet['C' + str(row + 8)].border = thin_border_bottom_left
                sheet['D' + str(row + 8)].border = thin_border_bottom_left
                sheet['E' + str(row + 8)].border = thin_border_bottom_left
                sheet['F' + str(row + 8)].border = thin_border_bottom_left
                sheet['G' + str(row + 8)].border = thin_border_bottom_left
                sheet['H' + str(row + 8)].border = thin_border_bottom_left
                sheet['I' + str(row + 8)].border = thin_border_bottom_left
                sheet['J' + str(row + 8)].border = thin_border_bottom_left
                sheet['K' + str(row + 8)].border = thin_border_bottom_left_right

        sheet.row_dimensions[8].height = 33
        sheet.column_dimensions["A"].width = 5
        sheet.column_dimensions["B"].width = 14
        sheet.column_dimensions["C"].width = 17
        sheet.column_dimensions["D"].width = 17
        sheet.column_dimensions["E"].width = 13
        sheet.column_dimensions["F"].width = 18
        sheet.column_dimensions["G"].width = 8
        sheet.column_dimensions["H"].width = 9
        sheet.column_dimensions["I"].width = 7
        sheet.column_dimensions["J"].width = 7
        sheet.column_dimensions["K"].width = 36

        wb.save(self.path + self.file_name)
        return


class convertExcel(object):
    def __init__(self, excel_file=''):
        self.excel_file = excel_file
        self.wb = load_workbook(filename=self.excel_file)
        self.sheet = self.wb.worksheets[0]
        self.row_count = self.sheet.max_row
        self.column_count = self.sheet.max_column

    def chunks(self, l, n):
        """Yield successive n-sized chunks from l."""
        for i in range(0, len(l), n):
            yield l[i:i + n]

    def array_data(self, header=True):
        excel_temp = []
        for row in self.sheet.iter_rows():
            for cell in row:
                if cell.row >= 9:
                    excel_temp.append(cell.internal_value)
                else:
                    pass
        chunk_data = list(self.chunks(excel_temp, 11))
        if header == True:
            final = []
            excel_temp_header = ['No', 'Point ID', 'Easting (X)', 'Northing (Y)',
                                 'Distance', 'Shared Vertices', 'Type', 'Polygon No.',
                                 'In/Out', 'Point Order', 'No_SK', 'IDP']
            final.append(excel_temp_header)
            for row in chunk_data:
                final.append([int(row[0]), str(row[1]), float(row[2]), float(row[3]),
                              float(row[4]), str(row[5]), str(row[6]), int(row[7]),
                              str(row[8]), int(row[9]), str(row[10]), str(row[7]) + str(row[8])])
        else:
            final = []
            excel_temp_header = []
            for row in chunk_data:
                final.append([int(row[0]), str(row[1]), float(row[2]), float(row[3]),
                              float(row[4]), str(row[5]), str(row[6]), int(row[7]),
                              str(row[8]), int(row[9]), str(row[10]), str(row[7]) + str(row[8])])

        final_sort = sorted(final, key=lambda x: (int(x[7]), str(x[8]), int(x[9])))
        return final_sort

    def array_info(self):
        lb_id = self.sheet['C2'].value
        lb_pt = self.sheet['C3'].value
        lb_ha = round(float(self.sheet['C4'].value.replace(",", "").replace(" ha", "")), 2)
        lb_date = self.sheet['C5'].value
        lb_proj = self.sheet['C6'].value
        lb_wkid = lb_proj[lb_proj.index("(") + 1:lb_proj.rindex(")")]
        return [lb_id, lb_pt, lb_ha,
                lb_date, lb_proj, lb_wkid]


class arrayToPolygon(object):
    def __init__(self, array_polygon, array_info, out_put_path=""):
        self.array_polygon = array_polygon
        self.array_info = array_info
        self.out_put_path = out_put_path + "\\"
        self.lb_id = self.array_info[0]
        self.lb_pt = self.array_info[1]
        self.lb_pt_name = (self.lb_pt.replace(' ', '_').replace('.', ''))
        self.lb_ha = self.array_info[2]
        self.lb_date = self.array_info[3]
        self.lb_proj = self.array_info[4]
        self.lb_wkid = self.array_info[5]
        self.point_name = self.lb_id[:-1] + "1"
        self.polygon_name = self.lb_id

    def second_smallest(self, numbers):
        m1, m2 = float('inf'), float('inf')
        for x in numbers:
            if x <= m1:
                m1, m2 = x, m1
            elif x < m2:
                m2 = x
        return m2

    def create_polygon(self):
        arcpy.Delete_management("in_memory")
        file_loc = []
        in_file = []
        out_file = []
        in_file_pt = []
        out_file_pt = []
        vertices_t = []
        vertices_mm = []
        point = arcpy.Point()
        array = arcpy.Array()
        geometry_type = "POLYGON"
        template = "#"
        has_m = "DISABLED"
        has_z = "DISABLED"
        spatial_reference = arcpy.SpatialReference(int(self.lb_wkid))

        if arcpy.Exists(self.out_put_path + self.polygon_name + ".shp"):
            arcpy.Delete_management(self.out_put_path + self.polygon_name + ".shp")
        else:
            pass
        if arcpy.Exists(self.out_put_path + self.point_name + ".shp"):
            arcpy.Delete_management(self.out_put_path + self.point_name + ".shp")
        else:
            pass

        for key, rows in groupby(self.array_polygon, lambda x: (x[11])):
            ioo = ''.join([i for i in key if not i.isdigit()])
            if ioo == 'in':
                iot = "I"
            elif ioo == 'out':
                iot = "O"

            pon = int(key.replace(ioo, ''))

            file_loc.append([self.out_put_path + key + ".shp", iot])
            if iot == "I":
                in_file.append(self.out_put_path + key + ".shp")
                in_file_pt.append(self.out_put_path + key + "_pt.shp")
            else:
                out_file.append(self.out_put_path + key + ".shp")
                out_file_pt.append(self.out_put_path + key + "_pt.shp")

            ft = arcpy.CreateFeatureclass_management(self.out_put_path, key + ".shp", geometry_type, template, has_m,
                                                     has_z, spatial_reference)

            arcpy.AddField_management(self.out_put_path + key + ".shp", "Polygon_No", "TEXT", field_length=21)
            arcpy.AddField_management(self.out_put_path + key + ".shp", "Company", "TEXT", field_length=31)
            arcpy.AddField_management(self.out_put_path + key + ".shp", "Vertices", "TEXT", field_length=5)
            arcpy.AddField_management(self.out_put_path + key + ".shp", "DateTime", "DATE")

            pt = arcpy.CreateFeatureclass_management(self.out_put_path, key + "_pt.shp", "POINT", template, has_m,
                                                     has_z, spatial_reference)

            arcpy.AddField_management(self.out_put_path + key + "_pt.shp", "Polygon_No", "TEXT", field_length=21)
            arcpy.AddField_management(self.out_put_path + key + "_pt.shp", "Point_ID", "TEXT", field_length=15)
            arcpy.AddField_management(self.out_put_path + key + "_pt.shp", "Company", "TEXT", field_length=31)
            arcpy.AddField_management(self.out_put_path + key + "_pt.shp", "DateTime", "DATE")

            icursor = arcpy.da.InsertCursor(ft, ["SHAPE@"])
            ptcursor = arcpy.da.InsertCursor(pt, ['Polygon_No', 'Point_ID', 'Company', 'DateTime', "SHAPE@"])
            total = 0
            ds = []
            rv = []
            for row in rows:
                ptcursor.insertRow((pon, row[1],
                                    self.lb_pt, datetime.datetime.now(),
                                    (row[2], row[3])))
                total += 1
                ds.append(row[4])
                point.X = row[2]
                point.Y = row[3]
                array.add(point)

            array.add(array.getObject(0))
            polygon = arcpy.Polygon(array)
            icursor.insertRow([polygon])
            array.removeAll()

            del icursor
            del ptcursor

            with arcpy.da.UpdateCursor(self.out_put_path + key + ".shp",
                                       ["Polygon_No", "Company", "Vertices", "DateTime"]) as cursor:
                for row in cursor:
                    row[0] = pon
                    row[1] = self.lb_pt
                    row[2] = total
                    row[3] = datetime.datetime.now()
                    cursor.updateRow(row)
            del cursor

            vertices_t.append([pon, total])
            vertices_mm.append([pon, max(ds), self.second_smallest(ds)])
        all_pt = out_file_pt + in_file_pt
        if len(in_file) > 0 and len(out_file) > 0:
            arcpy.Merge_management(in_file, self.out_put_path + "I.shp")
            arcpy.Merge_management(out_file, self.out_put_path + "O.shp")
            arcpy.Merge_management(all_pt, self.out_put_path + self.point_name + ".shp")
            arcpy.Erase_analysis(self.out_put_path + "I.shp", self.out_put_path + "O.shp",
                                 self.out_put_path + self.polygon_name + ".shp")
            for deleted in out_file:
                arcpy.Delete_management(deleted)
            for deleted in in_file:
                arcpy.Delete_management(deleted)

            for deleted in out_file_pt:
                arcpy.Delete_management(deleted)
            for deleted in in_file_pt:
                arcpy.Delete_management(deleted)

        elif len(in_file) > 0 and len(out_file) < 1:
            arcpy.Merge_management(in_file, self.out_put_path + self.polygon_name + ".shp")
            arcpy.Merge_management(in_file_pt, self.out_put_path + self.point_name + ".shp")
            for deleted in in_file:
                arcpy.Delete_management(deleted)
            for deleted in in_file_pt:
                arcpy.Delete_management(deleted)
        else:
            pass

        if arcpy.Exists(self.out_put_path + "I.shp"):
            arcpy.Delete_management(self.out_put_path + "I.shp")
        else:
            pass
        if arcpy.Exists(self.out_put_path + "O.shp"):
            arcpy.Delete_management(self.out_put_path + "O.shp")
        else:
            pass

        my_list2 = []
        for i, g in groupby(sorted(vertices_t), key=lambda x: x[0]):
            my_list2.append([i, sum(v[1] for v in g)])
        return [my_list2, vertices_mm]


class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        canvas.Canvas.__init__(self, *args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        """add page info to each page (page x of y)"""
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_number(num_pages)
            canvas.Canvas.showPage(self)
        canvas.Canvas.save(self)

    def draw_page_number(self, page_count):
        self.setFont("Helvetica", 6)
        self.line(25 * mm, 25 * mm, 185 * mm, 25 * mm)
        self.drawRightString(185 * mm, 20 * mm,
                             "Page %d of %d" % (self._pageNumber, page_count))


class createReport:
    def __init__(self, out_path="", data_array=[], info_array=[], detil_array=[]):
        if not detil_array[0]:
            test = ''
        else:
            test = detil_array[0]
        self.out_path = out_path + "\\"
        self.data_array = data_array
        self.info_array_vert = info_array[0]
        self.info_array_min = info_array[1]
        self.total_vertices = sum(a[1] for a in self.info_array_vert)
        self.detil_array = detil_array
        self.width, self.height = A4
        self.lb_id = test
        self.lb_pt = self.detil_array[1]
        self.lb_pt_name = (self.lb_pt.replace(' ', '_').replace('.', ''))
        self.lb_ha = '{:,.2f}'.format(self.detil_array[2])
        self.lb_date = self.detil_array[3].strftime("%d-%m-%Y")
        self.lb_proj = self.detil_array[4]
        self.lb_wkid = self.detil_array[5]
        self.styles = getSampleStyleSheet()

    def stylesheet(self):
        pads = 0
        return {
            'default': ParagraphStyle('default',
                                      fontName='Helvetica',
                                      fontSize=5,
                                      leading=6,
                                      leftIndent=0,
                                      rightIndent=0,
                                      firstLineIndent=0,
                                      alignment=TA_LEFT,
                                      spaceBefore=0,
                                      spaceAfter=0,
                                      bulletFontName='Helvetica',
                                      bulletFontSize=5,
                                      bulletIndent=0,
                                      textColor=black,
                                      backColor=None,
                                      wordWrap=None,
                                      borderWidth=0,
                                      borderPadding=0,
                                      borderColor=None,
                                      borderRadius=None,
                                      allowWidows=1,
                                      allowOrphans=0,
                                      textTransform=None,  # 'uppercase' | 'lowercase' | None
                                      endDots=None,
                                      splitLongWords=1, ),
            'table_default': TableStyle(
                [
                    ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
                    ('FONTSIZE', (0, 0), (-1, -1), 6),
                    ('LEADING', (0, 0), (-1, -1), 6),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    ('ALIGN', (2, 1), (2, -1), 'RIGHT'),
                    ('ALIGN', (3, 1), (3, -1), 'RIGHT'),
                    ('ALIGN', (4, 1), (4, -1), 'RIGHT'),
                    ('INNERGRID', (0, 0), (-1, -1), 0.1, colors.black),
                    ('BOX', (0, 0), (-1, -1), 0.1, colors.black),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                    ("LEFTPADDING", (0, 0), (-1, -1), 3),
                ]
            ),
            'table_info': TableStyle(
                [
                    ('FONT', (0, 0), (-1, -1), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 6),
                    ('LEADING', (0, 0), (-1, -1), 7),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
                    ('TOPPADDING', (0, 0), (-1, -1), 1),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 1),
                    ('LEFTPADDING', (0, 0), (-1, -1), 1),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ]
            ),
            'table_end_note': TableStyle(
                [
                    ('FONT', (0, 0), (-1, 0), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('LEADING', (0, 0), (-1, -1), 10),
                    ('FONTSIZE', (3, 0), (3, 0), 5),
                    ('LEADING', (3, 0), (3, 0), 4),
                    ('VALIGN', (3, 0), (3, 0), 'TOP'),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), pads),
                    ("TOPPADDING", (0, 0), (-1, -1), pads),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 1),
                    ("LEFTPADDING", (0, 0), (-1, -1), pads),
                ]
            ),
        }

    def build_flowabel_resume(self):
        data_to_show = [['No', 'Point ID', 'Easting (X)', 'Northing (Y)',
                         'Distance', "Shared\nVertices", 'Type',
                         'Polygon\nNo.', 'In/Out', 'Point\nOrder', 'Nomor SK']]
        for d in self.data_array:
            data_to_show.append([
                d[0], d[1], '{:,.2f}'.format(d[2]), '{:,.2f}'.format(d[3]),
                '{:,.2f}'.format(d[4]), d[5], d[6],
                d[7], d[8], d[9], d[10]
            ])
        nrows = len(data_to_show)
        rowHeights = nrows * [None]
        rowHeights[0] = .75 * cm
        information = [
            ['Landbank ID ', ':', self.lb_id, ' ', ' ', 'Jumlah Polygon ', ':', self.info_array_vert[-1][0]],
            ['Luas ', ':', self.lb_ha + ' ha', ' ', ' ', 'Jumlah Vertex ', ':', self.total_vertices],
            ['Tanggal ', ':', self.lb_date],
            ['Proyeksi ', ':', self.lb_proj]
        ]
        return [
            Paragraph("LAPORAN", ParagraphStyle('MyTitle',
                                                fontName='Helvetica-Bold',
                                                fontSize=16,
                                                leading=18,
                                                alignment=TA_LEFT,
                                                textColor=black)),
            Paragraph("Landbank PT. " + self.lb_pt.replace("PT.", '').title(), ParagraphStyle('MyTitle',
                                                                                              fontName='Helvetica-Bold',
                                                                                              alignment=TA_LEFT,
                                                                                              fontSize=14,
                                                                                              leading=15,
                                                                                              textColor=black)),
            HRFlowable(width="100%",
                       color=black,
                       lineCap='square',
                       spaceBefore=0.25 * cm,
                       spaceAfter=0,
                       thickness=1.5),
            HRFlowable(width="100%",
                       color=black,
                       lineCap='square',
                       spaceBefore=0.05 * cm,
                       spaceAfter=0,
                       thickness=2.5),
            Spacer(0, 0.25 * cm),

            Table(
                information, hAlign='LEFT',
                style=self.stylesheet()['table_info']
            ),

            Spacer(0, 0.35 * cm),

            Table(
                data_to_show, hAlign='LEFT',
                repeatRows=1,
                style=self.stylesheet()['table_default'], rowHeights=rowHeights
            ),

            Spacer(0, 0.45 * cm),

            Spacer(0, 0.25 * cm),

            Spacer(0, 0.15 * cm),

        ]

    def build_earthwork_resume(self):
        doc = BaseDocTemplate(self.out_path + self.lb_id + " (Report).pdf",
                              pagesize=A4,
                              pageTemplates=[],
                              showBoundary=0,
                              leftMargin=2.54 * cm,
                              rightMargin=2.54 * cm,
                              topMargin=2.54 * cm,
                              bottomMargin=2.54 * cm,
                              allowSplitting=1,
                              title="Landbank Report",
                              author="GIS")
        doc.addPageTemplates(
            [
                PageTemplate(
                    frames=[
                        Frame(
                            doc.leftMargin,
                            doc.bottomMargin,
                            doc.width,
                            doc.height,
                            id='left',
                            leftPadding=0.2 * cm,
                            rightPadding=0.2 * cm,
                            showBoundary=0  # set to 1 for debugging
                        ),
                    ]
                ),
            ]
        )
        doc.build(self.build_flowabel_resume(),
                  canvasmaker=NumberedCanvas)
        return "success build resume"


'''
excel = convertExcel(excel_file="D:\\AMS\\PYTHON\\PROJECT\\AUTOMATION\\LANBANK_IN_GIT\\Install\\AJP.xlsx")

info = excel.array_info()
data = excel.array_data(header=False)


cp = arrayToPolygon(array_polygon=data,
                    array_info=info,
                    out_put_path="D:\\AMS\\PYTHON\\PROJECT\\AUTOMATION\\LANBANK_IN_GIT\\Install\\res\\")

ress = cp.create_polygon()

pdf = createReport(out_path="D:\\AMS\\PYTHON\\PROJECT\\AUTOMATION\\LANBANK_IN_GIT\\Install\\res\\",
                   data_array=data,
                   info_array=ress,
                   detil_array=info)
print pdf.build_earthwork_resume()
'''
conv = convertVertices(feature_class=fc,
                       legal_type="",
                       header=True,
                       nama_pt='',
                       rounding=2)

array_result = conv.explode_polygon()
array_info = conv.information()
exc = exportExcel(path="D:\\AMS\\PYTHON\\PROJECT\\AUTOMATION\\LANBANK_IN_GIT\\Install\\",
                  file_name="PTWW",
                  info_array=array_info,
                  data_array=array_result
                  )
exc.createFile()
